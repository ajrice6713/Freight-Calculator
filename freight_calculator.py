# !/usr/bin/env python
#
#  name: freight_calculator.py
#  description: Simple freight rate calculator that uses DHL rates from an excel sheet
#  author: Andrew Rice
#  date: 6/28/19
#  version: 1.0
#  usage: python
#  notes:
#  python_version: 3.7.3
#  Copyright (c) 2019. Andrew Rice, All rights reserved
#  Last Modified: 7/3/19, 1:11 PM.


import openpyxl
import sys
import warnings
import time
import os
from datetime import datetime
import getpass
import socket


def adjust(weight):
    """Rounds the calculated weight up to the nearest 1/2 KG"""
    dec = float(weight) - int(weight)
    if (dec < .5) and (dec > 0):
        dec = .5
    elif dec >= .5:
        dec = 1
    else:
        dec = 0
    weight = int(weight) + dec
    return weight


def buffer(weight):    # Add a buffer to the calculated weights
    """Adds a buffer to the weight to give cushion when estimating"""
    buff = {    # Key = weight; Value = buffer to add
        5: 1,
        10: 2,
        25: 2.5,
        50: 5,
        75: 8,
        100: 12,
        150: 17,
        200: 20,
        500: 35
    }
    for wt, bf in buff.items():
        if weight <= wt:
            return weight + bf
    else:
        return weight + 40


def generate_cost(wt, col):
    """Iterate through the wt list to calculate a cost using the DHL rate sheet"""
    wt = float(wt)
    if wt <= 70:
        for i in range(15, 155, 1):
            if wt == float(imp_rates_ws.cell(row=i, column=1).value):
                wt = float(imp_rates_ws.cell(row=i, column=col).value)
    elif wt <= 150:
        wt *= float(imp_rates_ws.cell(row=158, column=col).value)
    elif wt <= 300:
        wt *= float(imp_rates_ws.cell(row=159, column=col).value)
    elif wt <= 999:
        wt *= float(imp_rates_ws.cell(row=160, column=col).value)
    else:
        wt *= float(imp_rates_ws.cell(row=161, column=col).value)
    return wt


# Get inbound country from user
# China - Zone G
# Taiwan, Hong Kong -  Zone H
# Thailand - Zone I
while True:
    zone = input('\nWhat country will this shipment be coming from?\n\nType \'q\' to quit\n'
                 'Press 1 for China\nPress 2 for Taiwan or Hong Kong\nPress 3 for Thailand\nCountry: ')
    if zone == 'q':
        sys.exit()

    try:    # restart loop if a valid int or 'q' isn't entered
        zone = int(zone)
    except ValueError:
        print('\nPlease enter a valid character\n')
        continue

    if zone < 1 or zone > 3:
        print('\nPlease enter 1, 2, or 3\n')
    else:
        break

if zone == 1:
    col = 9
elif zone == 2:
    col = 10
elif zone == 3:
    col = 11
else:
    print('There was an error processing your zone, please try again')
    time.sleep(1.5)
    sys.exit()

# Get quantities from user
quantity_list = input('\nEnter the quantities you would like to quote, separated by a space:\n')
quantity_list = list(map(int, quantity_list.split()))
quantity_list.sort()

# Get weight of 1 pc (in grams)
weight_g = float(input('\nEnter the weight of 1pc (in grams):\n'))
print('\nProcessing Quote\n')
weight_KG = weight_g/1000   # convert to KG (/1000)
weight_list = []
for i in range(len(quantity_list)):    # multiply weight by each quantity
    weight_list.append(weight_KG * quantity_list[i])
    weight_list[i] = adjust(weight_list[i])
print('Raw Weight:', weight_list)

# Buffer the weight via a scale
# ex. 1-10 KG, add 2 KG, 10.5-20 KG, add 3 KG, etc
buff_weight_list = []
for i in range(len(weight_list)):
    buff_weight_list.append(buffer(weight_list[i]))
print('Buffered Weight:', buff_weight_list, '\n')

# Match the qty and wt to the DHL rate sheet and return a variable
with warnings.catch_warnings():    # .wmf image in the excel file causes a warning
    warnings.simplefilter('ignore')
    wb = openpyxl.load_workbook('Z:\\AJ\\FREIGHT\\2019 FREIGHT RATES\\2019_DHL_RATES.xlsx', data_only=True)
    imp_rates_ws = wb['US Import Rates']    # import rates
freight_cost_raw = []
freight_cost_buffer = []
for i in range(len(weight_list)):
    freight_cost_raw.append(generate_cost(weight_list[i], col))
# print(freight_cost_raw)
for i in range(len(buff_weight_list)):
    freight_cost_buffer.append(generate_cost(buff_weight_list[i], col))
# print(freight_cost_buffer)

# Generate printout
print('Report Complete'.center(32, '-'))
print('Raw Cost:')
for i in range(len(weight_list)):
    print(str(quantity_list[i]) + ' pcs: $' +
          str(round(freight_cost_raw[i], 2)).rjust(25-len(str(quantity_list[i]))))
print('\nBuffered Cost:')
for i in range(len(buff_weight_list)):
    print(str(quantity_list[i]) + ' pcs: $' +
          str(round(freight_cost_buffer[i], 2)).rjust(25-len(str(quantity_list[i]))))
print('-' * 32)

# TODO: Generate a report that saves to a file on the server of every quote generated, and by whom
account = getpass.getuser()
host = socket.gethostname()
date = datetime.today().strftime('%m.%d.%Y')

# Generate a spreadsheet of the rate quotes
print('\nWould you like to generate an excel report of this data?')
ans = input('Press \'y\' to print a report, otherwise press any key to exit.\n')
if ans == 'y':
    print('\n Saving a report to your desktop')
    desktop = os.path.join(os.path.join(os.path.expanduser('~')), 'Desktop')
    report_wb = openpyxl.Workbook()
    report_ws = report_wb.active
    report_ws['A1'] = 'Freight Estimate'
    report_ws['A2'] = datetime.today().strftime('%m/%d/%Y')
    report_ws['A4'] = 'Quantity'
    report_ws['B4'] = 'Raw Cost'
    report_ws['C4'] = 'Buffered Cost'
    a = 0
    for x in range(5, len(quantity_list) + 5, 1):
        report_ws.cell(row=x, column=1).value = int(quantity_list[a])
        report_ws.cell(row=x, column=2).value = float(freight_cost_raw[a])
        report_ws.cell(row=x, column=3).value = float(freight_cost_buffer[a])
        a += 1
    report_wb.save(desktop + '\\freight_estimate.xlsx')
else:
    print('\nClosing...')
    time.sleep(.5)
    sys.exit()
